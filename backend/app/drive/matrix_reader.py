"""Reads the Archiving Matrix index, regardless of whether it lives as a
native Google Sheet or an .xlsx file uploaded into Drive. Returns a list of
row dicts keyed by the matrix's own header row (e.g. "Contract Number",
"Contractor", "Scope", "Contract Date" for 01-Contracts, or "Section ID",
"Section Title", "Description" for 02-Specifications) - we don't hardcode
column names here since the exact header wording is whatever the user's
real matrix uses.
"""

from __future__ import annotations

import io

from openpyxl import load_workbook

from app.drive import client
from app.drive.client import SHEET_MIME


def read_archiving_matrix(file_id: str, sheet_range: str = "A:Z") -> list[dict]:
    metadata = client.get_file_metadata(file_id)

    if metadata["mimeType"] == SHEET_MIME:
        rows = client.read_sheet_values(file_id, sheet_range)
    else:
        rows = _read_xlsx_rows(client.download_file(file_id))

    if not rows:
        return []

    header = [str(h).strip() for h in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(cell for cell in row):
            continue
        padded = list(row) + [""] * (len(header) - len(row))
        records.append(dict(zip(header, padded)))
    return records


def _read_xlsx_rows(file_bytes: bytes) -> list[list[str]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell) for cell in row])
    return rows
